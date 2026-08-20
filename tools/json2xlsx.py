#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
json2xlsx.py —— 把 EMS_VSG config/modbus/ 的 JSON 配置导出成 Excel 模板。

用法:
    python tools/json2xlsx.py [配置目录] [输出xlsx]

默认: 配置目录 = <仓库>/config/modbus, 输出 = <配置目录>/EMS_VSG配置模板.xlsx

生成的 Excel 含两个部分:
  - Sheet「设备配置」: 一行一个设备（devices.json）
  - 每个模板一个 Sheet: 顶部 sink.mode/base_addr + 点表（templates/*.json）

填表后用 tools/xlsx2json.py 转回 JSON。
"""
import os
import sys
import json

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import config_schema as sch


def _repo_root():
    return os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


HEADER_FONT = Font(bold=True)
HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
LABEL_FONT = Font(bold=True)
WRAP = Alignment(vertical="center")


def write_devices_sheet(ws, devices):
    cols = [c[0] for c in sch.DEVICE_COLUMNS]
    for prefix, keys, _types in sch.DEVICE_NESTED:
        for k in keys:
            cols.append(prefix + "_" + k)

    for j, col in enumerate(cols, start=1):
        c = ws.cell(1, j, col)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = WRAP

    for d in devices:
        row = []
        for f in sch.DEVICE_COLUMNS:
            row.append(d.get(f[0]))
        for prefix, keys, _types in sch.DEVICE_NESTED:
            obj = d.get(prefix)
            for k in keys:
                row.append(obj.get(k) if isinstance(obj, dict) else None)
        ws.append(row)

    ws.freeze_panes = "A2"
    for j, col in enumerate(cols, start=1):
        width = max(8, min(28, len(col) + 2))
        ws.column_dimensions[get_column_letter(j)].width = width


def write_template_sheet(ws, tname, tpl):
    # 模板 Sheet：纯读点表，从第 1 行直接是表头（sink 已独立成「写库映射」Sheet）
    headers = [c[0] for c in sch.POINT_COLUMNS]
    for j, h in enumerate(headers, start=1):
        c = ws.cell(sch.POINT_HEADER_ROW, j, h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        ws.column_dimensions[get_column_letter(j)].width = 14

    r = sch.POINT_START_ROW
    for p in tpl.get("read", {}).get("fields", []):
        for j, f in enumerate(sch.POINT_COLUMNS, start=1):
            ws.cell(r, j, p.get(f[0]))
        r += 1

    ws.freeze_panes = "A%d" % sch.POINT_START_ROW


def _write_flat_sheet(ws, headers, rows, widths):
    for j, h in enumerate(headers, start=1):
        c = ws.cell(1, j, h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        ws.column_dimensions[get_column_letter(j)].width = widths[j - 1]
    for row in rows:
        ws.append(row)
    ws.freeze_panes = "A2"


def write_sink_sheets(wb, cfg_dir):
    """把 config/modbus/sink/*.json 平铺成「写库映射」「写库表路由」两个 Sheet。"""
    sink_dir = os.path.join(cfg_dir, "sink")
    sink_files = sorted(n for n in os.listdir(sink_dir) if n.endswith(".json"))
    map_rows, route_rows = [], []
    for fn in sink_files:
        tkey = fn[:-5]
        with open(os.path.join(sink_dir, fn), encoding="utf-8") as f:
            sink = json.load(f)
        for mname, entries in (sink.get("field_maps", {}) or {}).items():
            for e in entries or []:
                map_rows.append((tkey, mname, e.get("name"), e.get("addr")))
        for t in sink.get("tables", []) or []:
            route_rows.append((tkey, t.get("field_map"), t.get("mysql"), t.get("mysql_prefix")))

    ws1 = wb.create_sheet(sch.SINK_SHEET)
    _write_flat_sheet(ws1, [c[0] for c in sch.SINK_COLUMNS], map_rows, [12, 12, 28, 8])
    ws2 = wb.create_sheet(sch.SINK_ROUTE_SHEET)
    _write_flat_sheet(ws2, [c[0] for c in sch.SINK_ROUTE_COLUMNS], route_rows, [12, 12, 16, 16])
    return sink_files


def main():
    cfg_dir = sys.argv[1] if len(sys.argv) > 1 else os.path.join(_repo_root(), "config", "modbus")
    out = sys.argv[2] if len(sys.argv) > 2 else os.path.join(cfg_dir, "EMS_VSG配置模板.xlsx")

    dev_path = os.path.join(cfg_dir, "devices.json")
    tpl_dir = os.path.join(cfg_dir, "templates")
    with open(dev_path, encoding="utf-8") as f:
        devices = json.load(f)["devices"]

    wb = Workbook()
    ws = wb.active
    ws.title = sch.DEVICES_SHEET
    write_devices_sheet(ws, devices)

    tpl_names = sorted(n for n in os.listdir(tpl_dir) if n.endswith(".json"))
    for tname in tpl_names:
        with open(os.path.join(tpl_dir, tname), encoding="utf-8") as f:
            tpl = json.load(f)
        write_template_sheet(wb.create_sheet(tname[:-5]), tname[:-5], tpl)

    sink_names = write_sink_sheets(wb, cfg_dir)

    wb.save(out)
    total_points = sum(
        len(json.load(open(os.path.join(tpl_dir, n), encoding="utf-8")).get("read", {}).get("fields", []))
        for n in tpl_names
    )
    print("已生成 Excel 模板: %s" % out)
    print("  设备 Sheet: %d 台设备" % len(devices))
    print("  模板 Sheet: %d 个（共 %d 个采集点，纯读）" % (len(tpl_names), total_points))
    print("  %s / %s Sheet: %d 个写库映射" % (sch.SINK_SHEET, sch.SINK_ROUTE_SHEET, len(sink_names)))
    print("填完表后运行: python tools/xlsx2json.py <该xlsx> --dry-run 预览，确认后去掉 --dry-run 生成 JSON")


if __name__ == "__main__":
    main()
