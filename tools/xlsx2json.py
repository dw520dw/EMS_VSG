#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
xlsx2json.py —— 把填好的 Excel（json2xlsx.py 生成的模板）转回
EMS_VSG config/modbus/ 的 devices.json + templates/*.json。

用法:
    python tools/xlsx2json.py <Excel文件> [--out 输出目录] [--dry-run]

  --out     输出目录，默认 <仓库>/config/modbus（会覆盖现有配置，注意先 git 提交/备份）
  --dry-run 只校验和打印将生成的配置，不写文件

校验失败时不写任何文件（拒绝生成错误配置）。
"""
import os
import sys
import json
import argparse

from openpyxl import load_workbook

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import config_schema as sch


def _repo_root():
    return os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


errors = []
warnings = []


def _empty(raw):
    return raw is None or (isinstance(raw, str) and raw.strip() == "")


def parse_value(raw, typ, col):
    """返回 (值, ok)；空单元格返回 (None, False) 表示"不生成该字段"。"""
    if _empty(raw):
        return None, False
    if typ == "str":
        return str(raw).strip(), True
    if typ == "int":
        try:
            return int(float(str(raw).strip())), True
        except ValueError:
            errors.append("[%s] 需为整数，收到 %r" % (col, raw))
            return None, False
    if typ == "float":
        try:
            return float(str(raw).strip()), True
        except ValueError:
            errors.append("[%s] 需为数字，收到 %r" % (col, raw))
            return None, False
    if typ == "bool":
        if isinstance(raw, bool):
            return raw, True
        s = str(raw).strip().lower()
        if s in ("true", "1", "yes", "y", "是"):
            return True, True
        if s in ("false", "0", "no", "n", "否", ""):
            return False, True
        errors.append("[%s] 需为 true/false，收到 %r" % (col, raw))
        return None, False
    if typ == "addr":
        if isinstance(raw, (int, float)):
            return int(raw), True
        s = str(raw).strip()
        if s.lower().startswith("0x"):
            return s, True  # 保留 "0x.." 字符串形式
        try:
            return int(s), True
        except ValueError:
            errors.append("[%s] 地址需为十进制或 0x 十六进制，收到 %r" % (col, raw))
            return None, False
    return str(raw).strip(), True


def parse_enum(raw, allowed, col):
    s = str(raw).strip()
    if s not in allowed:
        errors.append("[%s] 非法值 %r，允许: %s" % (col, raw, ", ".join(allowed)))
        return None
    return s


# ---------------- 设备 ----------------
def load_devices(ws):
    headers = [c.value for c in ws[1]]
    colmap = []  # 每列: None | ("flat", field) | ("nested", prefix, key, typ)
    flat_by_name = {f[0]: f for f in sch.DEVICE_COLUMNS}
    nested_by_name = {}
    for prefix, keys, types in sch.DEVICE_NESTED:
        for k in keys:
            nested_by_name["%s_%s" % (prefix, k)] = (prefix, k, types[k])

    for h in headers:
        if _empty(h):
            colmap.append(None)
            continue
        name = str(h).strip()
        if name in flat_by_name:
            colmap.append(("flat", flat_by_name[name]))
        elif name in nested_by_name:
            p, k, t = nested_by_name[name]
            colmap.append(("nested", p, k, t))
        else:
            warnings.append("[设备] 忽略未知列: %s" % name)
            colmap.append(None)

    devices = []
    for row in ws.iter_rows(min_row=2):
        d = {}
        nested = {}
        has_any = False
        for idx, cell in enumerate(row):
            if idx >= len(colmap) or colmap[idx] is None:
                continue
            col = colmap[idx]
            if col[0] == "flat":
                _, f = col
                v, ok = parse_value(cell.value, f[1], f[0])
                if ok and v is not None:
                    d[f[0]] = v
                    has_any = True
            else:
                _, prefix, key, typ = col
                v, ok = parse_value(cell.value, typ, "%s_%s" % (prefix, key))
                if ok and v is not None:
                    nested.setdefault(prefix, {})[key] = v
                    has_any = True
        if not has_any:
            continue
        for prefix, sub in nested.items():
            d[prefix] = sub
        devices.append(d)

    # 校验
    seen_ids = set()
    for d in devices:
        if "id" not in d:
            errors.append("[设备] 有一行缺少 id")
            continue
        if d["id"] in seen_ids:
            errors.append("[设备] id 重复: %s" % d["id"])
        seen_ids.add(d["id"])
        if "template" not in d:
            errors.append("[设备 %s] 缺少 template" % d["id"])
        if "transport" in d:
            if d["transport"] == "rtu" and "rtu_device" not in d:
                warnings.append("[设备 %s] rtu 设备未填 rtu_device，将使用默认串口" % d["id"])
            if d["transport"] == "tcp" and "tcp_ip" not in d:
                errors.append("[设备 %s] tcp 设备未填 tcp_ip" % d["id"])
    return devices


# ---------------- 模板（纯读点表） ----------------
def load_template(ws):
    fields = []
    seen_names = set()
    for row in ws.iter_rows(min_row=sch.POINT_START_ROW):
        vals = [c.value for c in row[:len(sch.POINT_COLUMNS)]]
        if all(_empty(v) for v in vals):
            continue
        p = {}
        for j, f in enumerate(sch.POINT_COLUMNS):
            v, ok = parse_value(vals[j], f[1], f[0])
            if ok and v is not None:
                p[f[0]] = v
        if "type" not in p:
            errors.append("[模板 %s] 有采集点缺 type" % ws.title)
            continue
        if "name" not in p:
            warnings.append("[模板 %s] 采集点缺 name（引擎会按 _f0/_f1 自动命名，若 C++ hook 按名取值需补全）" % ws.title)
        else:
            if p["name"] in seen_names:
                errors.append("[模板 %s] 采集点 name 重复: %s" % (ws.title, p["name"]))
            seen_names.add(p["name"])
        fields.append(p)

    return {"read": {"fields": fields}}


# ---------------- 写库映射（sink） ----------------
def load_sink_sheets(wb, templates):
    """读「写库映射」「写库表路由」两个 Sheet，重建 sink/<template>.json 并校验。"""
    sinks = {}

    def bucket(tpl):
        if tpl not in sinks:
            sinks[tpl] = {"field_maps": {}, "tables": []}
        return sinks[tpl]

    if sch.SINK_SHEET in wb.sheetnames:
        for row in wb[sch.SINK_SHEET].iter_rows(min_row=2, values_only=True):
            tpl, fmap, name, addr = (row + (None, None, None, None))[:4]
            if _empty(tpl):
                continue
            tpl = str(tpl).strip()
            if _empty(fmap) or _empty(name) or _empty(addr):
                errors.append("[写库映射] %s 行缺 template/field_map/name/addr" % tpl)
                continue
            fmap = str(fmap).strip()
            name = str(name).strip()
            try:
                addr = int(float(str(addr).strip()))
            except ValueError:
                errors.append("[写库映射] %s/%s addr 非整数: %r" % (tpl, name, addr))
                continue
            bucket(tpl)["field_maps"].setdefault(fmap, []).append({"name": name, "addr": addr})
    else:
        errors.append("缺少写库映射 Sheet: %s" % sch.SINK_SHEET)

    if sch.SINK_ROUTE_SHEET in wb.sheetnames:
        for row in wb[sch.SINK_ROUTE_SHEET].iter_rows(min_row=2, values_only=True):
            tpl, fmap, mysql, prefix = (row + (None, None, None, None))[:4]
            if _empty(tpl):
                continue
            tpl = str(tpl).strip()
            if _empty(fmap):
                errors.append("[写库表路由] %s 行缺 field_map" % tpl)
                continue
            entry = {"field_map": str(fmap).strip()}
            if not _empty(mysql):
                entry["mysql"] = str(mysql).strip()
            if not _empty(prefix):
                entry["mysql_prefix"] = str(prefix).strip()
            if "mysql" not in entry and "mysql_prefix" not in entry:
                errors.append("[写库表路由] %s 行缺 mysql/mysql_prefix" % tpl)
                continue
            bucket(tpl)["tables"].append(entry)
    else:
        errors.append("缺少写库表路由 Sheet: %s" % sch.SINK_ROUTE_SHEET)

    for tpl, s in sinks.items():
        if tpl not in templates:
            errors.append("[写库映射] 引用了不存在的模板 Sheet: %s" % tpl)
            continue
        known = {f.get("name") for f in templates[tpl]["read"]["fields"] if f.get("name")}
        for fmap, entries in s["field_maps"].items():
            for e in entries:
                if e["name"] not in known:
                    errors.append("[写库映射] %s/%s 不在模板 read.fields 中" % (tpl, e["name"]))
        for t in s["tables"]:
            if t["field_map"] not in s["field_maps"]:
                errors.append("[写库表路由] %s 引用了不存在的 field_map: %s" % (tpl, t["field_map"]))
    return sinks


# ---------------- 主流程 ----------------
def main():
    ap = argparse.ArgumentParser(description="Excel → EMS_VSG JSON 配置")
    ap.add_argument("xlsx", help="Excel 文件路径")
    ap.add_argument("--out", default=None, help="输出目录，默认 config/modbus")
    ap.add_argument("--dry-run", action="store_true", help="只校验/预览，不写文件")
    args = ap.parse_args()

    out_dir = args.out or os.path.join(_repo_root(), "config", "modbus")

    wb = load_workbook(args.xlsx, data_only=True)
    if sch.DEVICES_SHEET not in wb.sheetnames:
        errors.append("缺少设备 Sheet: %s" % sch.DEVICES_SHEET)

    devices = load_devices(wb[sch.DEVICES_SHEET])
    templates = {}
    for sheet_name in wb.sheetnames:
        if sheet_name in (sch.DEVICES_SHEET, sch.SINK_SHEET, sch.SINK_ROUTE_SHEET):
            continue
        templates[sheet_name] = load_template(wb[sheet_name])
    sinks = load_sink_sheets(wb, templates)

    # 设备引用的模板是否都存在
    used_templates = {d.get("template") for d in devices if "template" in d}
    for t in used_templates:
        if t not in templates:
            errors.append("[设备] 引用了不存在的模板 Sheet: %s" % t)

    for w in warnings:
        print("[警告] %s" % w)
    if errors:
        print("\n发现 %d 处错误，已拒绝生成（未写任何文件）：" % len(errors))
        for e in errors:
            print("  [错误] %s" % e)
        sys.exit(1)

    print("校验通过。将生成:")
    print("  devices.json   %d 台设备: %s" % (len(devices), ", ".join(d.get("id") for d in devices)))
    for n, t in templates.items():
        print("  templates/%s.json  %d 个采集点" % (n, len(t["read"]["fields"])))
    for n, s in sinks.items():
        print("  sink/%s.json      %d 个写点" % (n, sum(len(e) for e in s["field_maps"].values())))

    if args.dry_run:
        print("\n[dry-run] 未写任何文件。确认无误后去掉 --dry-run 重新运行。")
        return

    os.makedirs(os.path.join(out_dir, "templates"), exist_ok=True)
    os.makedirs(os.path.join(out_dir, "sink"), exist_ok=True)
    with open(os.path.join(out_dir, "devices.json"), "w", encoding="utf-8") as f:
        f.write(json.dumps({"devices": devices}, ensure_ascii=False, indent=2) + "\n")
    for n, t in templates.items():
        with open(os.path.join(out_dir, "templates", n + ".json"), "w", encoding="utf-8") as f:
            f.write(sch.dump_config(t))
    for n, s in sinks.items():
        with open(os.path.join(out_dir, "sink", n + ".json"), "w", encoding="utf-8") as f:
            f.write(sch.dump_config(s))
    print("\n已写入: %s" % os.path.abspath(out_dir))
    print("注意: 目标板配置按 mtime 缓存、无热加载，改配置需重启 collect 进程。")


if __name__ == "__main__":
    main()
